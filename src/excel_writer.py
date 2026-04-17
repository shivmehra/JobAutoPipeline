# Excel writer with formatting for resume matching scores
import os
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelWriter:
    def __init__(self, file_path, sheet_name='Sheet1'):
        self.file_path = file_path
        self.sheet_name = sheet_name

        # Ensure the directory exists
        dir_path = os.path.dirname(file_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        else:
            print("Warning: No directory specified for Excel file.")

    def _organize_columns(self, df):
        """
        Reorganize columns to put match_score and rejection_reason at the front for visibility.
        Priority order: match_score, rejection_reason, job_title, company_name, ...rest
        """
        # Desired priority columns
        priority_cols = ['match_score', 'rejection_reason', 'job_title', 'company_name', 'job_type']
        
        # Get existing columns
        existing_cols = list(df.columns)
        
        # Build new column order
        new_cols = []
        for col in priority_cols:
            if col in existing_cols:
                new_cols.append(col)
        
        # Add remaining columns
        for col in existing_cols:
            if col not in new_cols:
                new_cols.append(col)
        
        return df[new_cols]

    def _apply_formatting(self, writer, sheet_name):
        """Apply Excel formatting (colors, borders, etc.)"""
        try:
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            score_high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            score_mid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow
            score_low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Find match_score column
            score_col_idx = None
            rejection_col_idx = None
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value == 'match_score':
                    score_col_idx = col_idx
                elif cell.value == 'rejection_reason':
                    rejection_col_idx = col_idx
            
            # Format data rows
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                for col_idx, cell in enumerate(row, 1):
                    cell.border = thin_border
                    
                    # Score column formatting
                    if col_idx == score_col_idx and cell.value is not None:
                        cell.alignment = center_alignment
                        try:
                            score = int(cell.value)
                            if score >= 7:
                                cell.fill = score_high_fill
                            elif score >= 4:
                                cell.fill = score_mid_fill
                            else:
                                cell.fill = score_low_fill
                        except:
                            pass
                    
                    # Rejection reason column
                    elif col_idx == rejection_col_idx:
                        cell.alignment = left_alignment
                    
                    # Other columns
                    else:
                        cell.alignment = left_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value or ''))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            print(f"Warning: Could not apply Excel formatting: {e}")

    def write_data(self, data, worksheet_name=None, sort_by_score=True):
        """
        Write data to Excel file with optional sorting by match_score.
        
        Args:
            data: List of dictionaries or DataFrame
            worksheet_name: Custom sheet name
            sort_by_score: Sort by match_score descending (best matches first)
        """
        try:
            if not data:
                print("No data to write")
                return False

            # Use provided worksheet_name or default
            sheet_name = worksheet_name or self.sheet_name

            # Convert data to DataFrame
            df = pd.DataFrame(data)
            
            # Sort by match_score if column exists and sort_by_score is True
            if sort_by_score and 'match_score' in df.columns:
                df = df.sort_values('match_score', ascending=False, na_position='last')
            
            # Organize columns for better visibility
            df = self._organize_columns(df)

            # Write to Excel
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting
                self._apply_formatting(writer, sheet_name)

            total_rows = len(df)
            high_score = len(df[df['match_score'] >= 7]) if 'match_score' in df.columns else 0
            
            print(f"✓ Data written to Excel file: {self.file_path}")
            print(f"  Total rows: {total_rows}")
            if 'match_score' in df.columns:
                print(f"  High matches (score ≥ 7): {high_score}")
            
            return True
        except Exception as e:
            print(f"✗ Error writing to Excel file: {e}")
            return False

    def append_data(self, data, worksheet_name=None):
        """Append data to existing Excel file"""
        try:
            if not data:
                print("No data to append")
                return False

            sheet_name = worksheet_name or self.sheet_name

            # Check if file exists
            if os.path.exists(self.file_path):
                # Read existing data
                try:
                    existing_df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                except:
                    # Sheet doesn't exist, create new
                    existing_df = pd.DataFrame()
            else:
                existing_df = pd.DataFrame()

            # Convert new data to DataFrame
            new_df = pd.DataFrame(data)

            # Append new data
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)

            # Write back to Excel with mode='a' to append instead of overwrite
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Data appended to Excel file: {self.file_path}")
            return True
        except Exception as e:
            print(f"Error appending to Excel file: {e}")
            return False